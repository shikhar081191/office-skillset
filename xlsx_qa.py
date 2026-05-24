"""Dependency-light structural quality checks for Excel deliverables."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


PLACEHOLDER_RE = re.compile(r"(lorem ipsum|\bplaceholder\b|\bTODO\b|\[\s*insert\b)", re.IGNORECASE)
FORMULA_ERROR_RE = re.compile(r"#(REF|DIV/0|VALUE|NAME|N/A|NULL|NUM)!?", re.IGNORECASE)


def _issue(level: str, sheet: str | None, check: str, message: str) -> dict:
    issue = {"level": level, "check": check, "message": message}
    if sheet is not None:
        issue["sheet"] = sheet
    return issue


def audit_xlsx(path: str | Path, final: bool = False) -> dict:
    xlsx_path = Path(path)
    errors = []
    warnings = []
    checks = {"sheets": 0, "cells": 0, "formulas": 0, "placeholder_text": 0, "layout": 0}
    if not xlsx_path.exists():
        errors.append(_issue("error", None, "file", f"File not found: {xlsx_path}"))
        return _report(xlsx_path, final, errors, warnings, checks)
    try:
        workbook = load_workbook(str(xlsx_path), data_only=False)
    except Exception as exc:
        errors.append(_issue("error", None, "file", f"Could not open XLSX: {exc}"))
        return _report(xlsx_path, final, errors, warnings, checks)

    checks["sheets"] = len(workbook.sheetnames)
    if not workbook.sheetnames:
        errors.append(_issue("error", None, "sheets", "Workbook contains no worksheets."))

    for worksheet in workbook.worksheets:
        populated = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                populated += 1
                checks["cells"] += 1
                text = str(cell.value)
                if text.startswith("="):
                    checks["formulas"] += 1
                    if FORMULA_ERROR_RE.search(text):
                        errors.append(_issue("error", worksheet.title, "formula_error", f"{cell.coordinate} contains an invalid formula reference."))
                if PLACEHOLDER_RE.search(text):
                    checks["placeholder_text"] += 1
                    issue = _issue(
                        "error" if final else "warning",
                        worksheet.title,
                        "placeholder_text",
                        f"{cell.coordinate} contains unfinished placeholder text.",
                    )
                    (errors if final else warnings).append(issue)
        if populated == 0:
            warnings.append(_issue("warning", worksheet.title, "empty_sheet", "Worksheet has no content."))
        for key, dimension in worksheet.column_dimensions.items():
            if dimension.width and dimension.width > 50:
                checks["layout"] += 1
                warnings.append(_issue("warning", worksheet.title, "layout", f"Column {key} is unusually wide ({dimension.width:g})."))

    return _report(xlsx_path, final, errors, warnings, checks)


def _report(path, final, errors, warnings, checks):
    return {
        "path": str(path),
        "final": final,
        "passed": not errors,
        "errors": errors,
        "warnings": warnings,
        "checks": checks,
    }


def ensure_xlsx_quality(path: str | Path, final: bool = False, report_path: str | Path | None = None) -> dict:
    report = audit_xlsx(path, final=final)
    if report_path:
        target = Path(report_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(json.dumps(report, indent=2), encoding="utf-8")
    if report["errors"]:
        messages = "; ".join(item["message"] for item in report["errors"])
        raise RuntimeError(f"XLSX quality gate failed: {messages}")
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit an XLSX before delivery.")
    parser.add_argument("xlsx_path")
    parser.add_argument("--final", action="store_true")
    parser.add_argument("--report", dest="report_path")
    args = parser.parse_args()
    report = audit_xlsx(args.xlsx_path, final=args.final)
    if args.report_path:
        Path(args.report_path).write_text(json.dumps(report, indent=2), encoding="utf-8")
    if report["errors"]:
        for issue in report["errors"]:
            print(f"ERROR: {issue['message']}")
        return 1
    print(f"QA passed: {len(report['warnings'])} warning(s), {args.xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
