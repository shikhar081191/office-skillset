"""Create .xlsx files using openpyxl (pure Python, no npm needed).

Covers: data entry, formulas, formatting, financial colour conventions,
tables, charts, and multi-sheet workbooks.

Usage:
    python create_xlsx.py          # runs demo
    from create_xlsx import XlsxBuilder
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


# ── financial colour conventions (from skill) ─────────────────────────────────
# Blue  = hardcoded inputs
# Black = formulas
# Green = cross-sheet links
# Red   = external links

COLOR_INPUT    = "FF0000FF"   # blue   - hardcoded user inputs (ARGB)
COLOR_FORMULA  = "FF000000"   # black  - calculated cells (ARGB)
COLOR_XSHEET   = "FF008000"   # green  - cross-sheet references (ARGB)
COLOR_EXTERNAL = "FFFF0000"   # red    - external file links (ARGB)
COLOR_HEADER   = "FF1E2761"   # header fill (ARGB)
COLOR_HEADER_FONT = "FFFFFFFF"
COLOR_ALT_ROW  = "FFEEF2FF"   # alternating row fill (ARGB)
COLOR_ASSUMPTION = "FFFFFF00" # yellow - key assumptions (ARGB)


class XlsxBuilder:
    """Fluent builder for Excel workbooks following financial model conventions."""

    def __init__(self):
        self.wb = Workbook()
        # Remove default sheet — we'll add explicitly
        self.wb.remove(self.wb.active)
        self._sheets = {}

    def add_sheet(self, name: str) -> "SheetBuilder":
        ws = self.wb.create_sheet(title=name)
        builder = SheetBuilder(ws, self)
        self._sheets[name] = builder
        return builder

    def save(self, output_path: str, validate=True, final=False, report_path=None) -> Path:
        """Save the workbook and run structural QA unless explicitly disabled."""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(path)
        if validate:
            from xlsx_qa import ensure_xlsx_quality
            ensure_xlsx_quality(path, final=final, report_path=report_path)
        return path


class SheetBuilder:
    """Builder for a single worksheet."""

    def __init__(self, ws, parent: XlsxBuilder):
        self.ws = ws
        self.parent = parent
        self._set_defaults()

    def _set_defaults(self):
        self.ws.sheet_view.showGridLines = True
        # Default column width
        self.ws.column_dimensions["A"].width = 22

    # ── data writing ─────────────────────────────────────────────────────────

    def headers(self, row: int, cols: list[str],
                start_col=1, col_width=14) -> "SheetBuilder":
        """Write a styled header row."""
        for i, header in enumerate(cols):
            col = start_col + i
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=COLOR_HEADER_FONT, name="Arial", size=11)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin_border()
            self.ws.column_dimensions[get_column_letter(col)].width = col_width
        self.ws.row_dimensions[row].height = 30
        return self

    def input_cell(self, row: int, col: int, value,
                   number_format=None, comment=None) -> "SheetBuilder":
        """Write a hardcoded input (blue text by convention)."""
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.font = Font(color=COLOR_INPUT, name="Arial", size=11)
        cell.border = _thin_border()
        if number_format:
            cell.number_format = number_format
        if comment:
            from openpyxl.comments import Comment
            cell.comment = Comment(comment, "Model")
        return self

    def formula_cell(self, row: int, col: int, formula: str,
                     number_format=None) -> "SheetBuilder":
        """Write a formula cell (black text by convention)."""
        cell = self.ws.cell(row=row, column=col, value=formula)
        cell.font = Font(color=COLOR_FORMULA, name="Arial", size=11)
        cell.border = _thin_border()
        if number_format:
            cell.number_format = number_format
        return self

    def label_cell(self, row: int, col: int, value: str,
                   bold=False) -> "SheetBuilder":
        """Write a label/description cell."""
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, name="Arial", size=11)
        cell.border = _thin_border()
        return self

    def assumption_cell(self, row: int, col: int, value,
                        number_format=None) -> "SheetBuilder":
        """Key assumption cell — yellow background."""
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.font = Font(color=COLOR_INPUT, name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_ASSUMPTION)
        cell.border = _thin_border()
        if number_format:
            cell.number_format = number_format
        return self

    def data_table(self, start_row: int, start_col: int,
                   headers: list, rows: list,
                   col_widths: list = None,
                   alt_rows=True) -> "SheetBuilder":
        """
        Write a formatted data table.
        rows: list of lists. Values starting with '=' are treated as formulas.
        """
        # Header
        self.headers(start_row, headers, start_col=start_col,
                     col_width=14 if not col_widths else col_widths[0])

        if col_widths:
            for i, w in enumerate(col_widths):
                self.ws.column_dimensions[get_column_letter(start_col + i)].width = w

        # Data rows
        for r_offset, row_data in enumerate(rows):
            r = start_row + 1 + r_offset
            fill = PatternFill("solid", fgColor=COLOR_ALT_ROW) if (alt_rows and r_offset % 2 == 1) else None

            for c_offset, value in enumerate(row_data):
                c = start_col + c_offset
                cell = self.ws.cell(row=r, column=c, value=value)
                cell.font = Font(
                    color=COLOR_FORMULA if isinstance(value, str) and value.startswith("=") else COLOR_INPUT,
                    name="Arial", size=11,
                )
                cell.border = _thin_border()
                if fill:
                    cell.fill = fill

        return self

    def freeze_panes(self, cell: str) -> "SheetBuilder":
        """Freeze rows/cols above and to the left of cell (e.g. 'B2')."""
        self.ws.freeze_panes = cell
        return self

    def col_width(self, col: int, width: float) -> "SheetBuilder":
        self.ws.column_dimensions[get_column_letter(col)].width = width
        return self

    def row_height(self, row: int, height: float) -> "SheetBuilder":
        self.ws.row_dimensions[row].height = height
        return self

    def add_bar_chart(self, title: str,
                      data_min_row: int, data_max_row: int,
                      data_min_col: int, data_max_col: int,
                      categories_col: int,
                      anchor_cell: str = "A20") -> "SheetBuilder":
        """Add a bar chart from a data range."""
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.style = 10
        chart.grouping = "clustered"
        chart.width = 15
        chart.height = 10

        data = Reference(self.ws,
                         min_col=data_min_col, max_col=data_max_col,
                         min_row=data_min_row, max_row=data_max_row)
        cats = Reference(self.ws,
                         min_col=categories_col,
                         min_row=data_min_row + 1, max_row=data_max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        self.ws.add_chart(chart, anchor_cell)
        return self

    def done(self) -> XlsxBuilder:
        """Return to the parent XlsxBuilder."""
        return self.parent


# ── shared helpers ─────────────────────────────────────────────────────────────

def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


# Number format constants (use with input_cell / formula_cell)
FMT_CURRENCY   = '$#,##0;($#,##0);"-"'   # dollars, negatives in parens, zeros as dash
FMT_CURRENCY_M = '$#,##0.0,,;($#,##0.0,,);"-"'  # millions
FMT_PCT        = "0.0%"
FMT_PCT2       = "0.00%"
FMT_MULTIPLE   = '0.0"x"'
FMT_INT        = "#,##0"
FMT_2DP        = "#,##0.00"


# ── demo ──────────────────────────────────────────────────────────────────────

def demo():
    b = XlsxBuilder()

    # ── Sheet 1: Assumptions ──────────────────────────────────────────────────
    s1 = b.add_sheet("Assumptions")
    s1.label_cell(1, 1, "MODEL ASSUMPTIONS", bold=True)
    s1.label_cell(3, 1, "Revenue Growth Rate", bold=False)
    s1.assumption_cell(3, 2, 0.08, FMT_PCT)          # yellow — key input
    s1.label_cell(4, 1, "EBITDA Margin")
    s1.assumption_cell(4, 2, 0.22, FMT_PCT)
    s1.label_cell(5, 1, "Tax Rate")
    s1.input_cell(5, 2, 0.25, FMT_PCT)               # blue — regular input
    s1.col_width(1, 24).col_width(2, 16)

    # ── Sheet 2: P&L ──────────────────────────────────────────────────────────
    s2 = b.add_sheet("P&L")
    s2.headers(1, ["", "FY2023A", "FY2024A", "FY2025E", "FY2026E", "FY2027E"],
               start_col=1, col_width=14)

    labels = ["Revenue", "Growth %", "", "EBITDA", "EBITDA Margin", "", "Net Income"]
    for i, label in enumerate(labels):
        s2.label_cell(2 + i, 1, label, bold=(label != ""))

    # Revenue — inputs for actuals, formulas for estimates
    s2.input_cell(2, 2, 1000, FMT_INT)   # FY23A
    s2.input_cell(2, 3, 1080, FMT_INT)   # FY24A
    # FY25E onward: formula using assumption
    for col in [4, 5, 6]:
        prev = get_column_letter(col - 1)
        s2.formula_cell(2, col, f"={prev}2*(1+Assumptions!$B$3)", FMT_INT)

    # Growth %
    for col in [3, 4, 5, 6]:
        prev = get_column_letter(col - 1)
        curr = get_column_letter(col)
        s2.formula_cell(3, col, f"=({curr}2-{prev}2)/{prev}2", FMT_PCT)

    # EBITDA
    for col in [2, 3, 4, 5, 6]:
        c = get_column_letter(col)
        s2.formula_cell(5, col, f"={c}2*Assumptions!$B$4", FMT_INT)

    # EBITDA Margin
    for col in [2, 3, 4, 5, 6]:
        c = get_column_letter(col)
        s2.formula_cell(6, col, f"={c}5/{c}2", FMT_PCT)

    # Net Income
    for col in [2, 3, 4, 5, 6]:
        c = get_column_letter(col)
        s2.formula_cell(8, col, f"={c}5*(1-Assumptions!$B$5)", FMT_INT)

    s2.freeze_panes("B2")
    s2.col_width(1, 18)

    # ── Sheet 3: Factor Returns ───────────────────────────────────────────────
    s3 = b.add_sheet("Factor Returns")
    s3.data_table(
        start_row=1, start_col=1,
        headers=["Factor", "Return (%)", "Vol (%)", "Sharpe", "Max DD (%)"],
        rows=[
            ["Equity Beta",    0.032,  0.141, "=C3/D3", -0.087],
            ["Credit Spread",  -0.018, 0.064, "=C4/D4", -0.031],
            ["Rates Duration", 0.009,  0.042, "=C5/D5", -0.018],
            ["FX Carry",       0.021,  0.087, "=C6/D6", -0.044],
            ["Momentum",       0.041,  0.112, "=C7/D7", -0.063],
        ],
        col_widths=[20, 14, 14, 14, 16],
    )
    # Format numeric columns
    for row in range(2, 7):
        for col in [2, 3, 5]:
            s3.ws.cell(row=row, column=col).number_format = FMT_PCT2
        s3.ws.cell(row=row, column=4).number_format = FMT_2DP

    out = b.save("demo_output.xlsx")
    print(f"Created: {out}")
    print("Sheets:", b.wb.sheetnames)


if __name__ == "__main__":
    demo()
